import os
from flask import Flask, request, jsonify, send_from_directory
import openpyxl
from datetime import (
    datetime,
)


BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")
STATIC_DIR = os.path.join(BASE_DIR, "static")
DB_DIR = os.path.join(
    os.path.dirname(__file__), "..", "db"
)
EXCEL_FILE = os.path.join(DB_DIR, "clientes.xlsx")
COLUMNS = [
    "ID",
    "Nome",
    "CPF",
    "Email",
    "Telefone",
    "Endereço",
    "Observações",
    "Data Cadastro",
]

def init_excel():
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR)

    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Clientes" 
        sheet.append(COLUMNS)
        workbook.save(EXCEL_FILE)


app = Flask(__name__, static_folder=STATIC_DIR, static_url_path="/" + STATIC_DIR)

@app.route("/")
def home():
    return send_from_directory(FRONTEND_DIR, "index.html")

@app.route("/consulta")
def consulta_page():
    return send_from_directory(FRONTEND_DIR, "consulta.html")

@app.route("/alterar")
def alterar_page():
    return send_from_directory(FRONTEND_DIR, "alterar.html")

@app.route("/assets/<path:filename>")
def assets(filename):
    return send_from_directory("../frontend/assets", filename)

@app.route("/cadastrar", methods=["POST"])
def cadastrar_cliente():
    try:
        data = request.json  
        required_fields = ["nome", "cpf", "email", "telefone", "endereco"]
        if not all(field in data and data[field] for field in required_fields):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Todos os campos obrigatórios devem ser preenchidos.",
                    }
                ),
                400,
            )
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        last_id = 0
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1).value or 0
        new_id = last_id + 1

        novo_cliente = [
            new_id,
            data.get("nome"),
            data.get("cpf"),
            data.get("email"),
            data.get("telefone"),
            data.get("endereco"),
            data.get("observacoes", ""),
            datetime.now().strftime("%Y-%m-%d"), 
        ]
        sheet.append(novo_cliente)
        workbook.save(EXCEL_FILE)
        return (
            jsonify(
                {
                    "status": "success",
                    "message": "Cliente cadastrado com sucesso!",
                    "id": new_id,
                }
            ),
            201,
        )
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao salvar no servidor: {e}"}),
            500,
        )

@app.route("/buscar", methods=["GET"])
def buscar_clientes():

    nome_query = request.args.get("nome", " ").lower()

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        resultados = []

        for row in sheet.iter.rows(main_row=2, values_only=True):
            cliente = dict(zip(COLUMNS, row))
            nome_cliente = (cliente.get("Nome") or "").lower()

            if nome_query in nome_cliente:
                resultados.append(cliente)

        return jsonify(resultados)

    except FileNotFoundError:
        return (
            jsonify({"status": "error", "message": "Arquivo de dados não encontrado"}),
            484,
        )
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao ler os dados: {e}"}),
            500,
        )

@app.route("/api/atualizar/<int:cliente_id>", methods=["POST"])
def atualizar_cliente(cliente_id):

    try:
        data = request.json
        
        workbook = openpyxl.load_workbooK(EXCEL_FILE)

        sheet = workbook.active

        row_to_update = -1

        for row_idx in range(2, sheet.max_row +1):
            if sheet.cell(row= row_idx, column=1).value == cliente_id:
                row_to_update = row_idx
                break

        if row_to_update ==-1:
            return(
                jsonify({
                    "status": "error",
                    "massage": "Cliente não encontrado para atualização.",
                }),
                404,
            )
        
        sheet.cell(row=row_to_update, column=2, value=data.get("nome"))
        sheet.cell(row=row_to_update, column=3, value=data.get("cpf"))
        sheet.cell(row=row_to_update, column=4, value=data.get("email"))
        sheet.cell(row=row_to_update, column=5, value=data.get("telefone"))
        sheet.cell(row=row_to_update, column=6, value=data.get("endereco"))
        sheet.cell(row=row_to_update, column=7, value=data.get("observacoes"))

        workbook.save(EXCEL_FILE)

        return jsonify({
            "status": "succes",
            "message": "Dados do cliente atualizados com sucesso!",
        })

    except Exception as e:

            return (
                jsonify({"status": "error", "message": f"Erro ao atualizar os dados:{e}"}),
                500,
            )




@app.route("/cliente/<int:cliente_id>", methods=["GET"])
def get_cliente(cliente_id):
    
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE).active
        sheet = workbook.active

        for row_idx in range(2, sheet.max_row + 1):
            row_id = sheet.cell(row=row_idx, column=1).value
            if row_id == cliente_id:
                row_values = [cell.value for cell in sheet[row_idx]]
                cliente = dict(zip(COLUMNS. row_values))
                return jsonify(cliente)

        return jsonify({"status": "error", "message": "Cliente não encontrado."}), 404
    
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao buscar o cliente: {e}"}),
            500,

        )



if __name__ == "__main__":
    print("BASE_DIR:", BASE_DIR)
    print("FRONTEND_DIR:", FRONTEND_DIR)
    print("STATIC_DIR:", STATIC_DIR)
    init_excel()
    app.run(debug=True)
